# =============================================
# generate_answer_sheet.py
# Generates an Excel answer sheet from sequences.json
# Conditions with the same trial number are grouped
# into one sheet, stacked vertically.
# Run: python generate_answer_sheet.py
# Requires: pip install openpyxl
# =============================================
import json
import re
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter

SEQ_FILE    = 'sequences.json'
OUTPUT_FILE = 'answer_sheet.xlsx'

# Display order of conditions within each trial sheet
CONDITION_ORDER = [
    'NC_1back', 'NC_2back', 'NC_3back',
    'VLTC_1back', 'VLTC_2back', 'VLTC_3back'
]

# Header color per condition (NC=blue, VLTC=green)
COND_COLORS = {
    'NC_1back':   '1A5276',
    'NC_2back':   '1F618D',
    'NC_3back':   '2874A6',
    'VLTC_1back': '117A65',
    'VLTC_2back': '1E8449',
    'VLTC_3back': '239B56',
}

# -----------------------------------------------
# Helpers
# -----------------------------------------------
def get_nback(key):
    m = re.search(r'(\d)back', key)
    return int(m.group(1)) if m else None

def get_correct(seq, i, n):
    if i < n:
        return None
    return seq[i] == seq[i - n]

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom_border():
    thin = Side(style='thin')
    thick = Side(style='medium')
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def header_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

# -----------------------------------------------
# Write one condition block starting at start_row.
# Returns the next available row after the block.
# -----------------------------------------------
def write_condition_block(ws, start_row, cond_key, seq):
    n = get_nback(cond_key)
    color = COND_COLORS.get(cond_key, '404040')
    valid = len(seq) - n
    col_widths = [4, 10, 12, 14, 10, 8]

    # --- Condition title ---
    ws.merge_cells(f'A{start_row}:F{start_row}')
    title_cell = ws[f'A{start_row}']
    title_cell.value = cond_key.replace('_', '  ')
    title_cell.font = Font(bold=True, size=11, color='FFFFFF')
    title_cell.fill = header_fill(color)
    title_cell.alignment = center()
    title_cell.border = thin_border()
    ws.row_dimensions[start_row].height = 18

    # --- Sub header ---
    sub_row = start_row + 1
    ws.merge_cells(f'A{sub_row}:F{sub_row}')
    sub = ws[f'A{sub_row}']
    sub.value = f'{n}-back   |   {len(seq)} stimuli   |   valid responses: {valid}'
    sub.font = Font(size=9, color='555555')
    sub.alignment = center()
    sub.border = thin_border()
    ws.row_dimensions[sub_row].height = 14

    # --- Column headers ---
    col_row = start_row + 2
    headers = ['#', 'Position', f'{n} step(s) ago', 'Correct answer', 'Response', 'O / X']
    header_colors = ['404040', '404040', '404040', '404040', 'AAAAAA', 'AAAAAA']

    for col, (h, w, c) in enumerate(zip(headers, col_widths, header_colors), start=1):
        cell = ws.cell(row=col_row, column=col, value=h)
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = header_fill(c)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[col_row].height = 16

    # --- Data rows ---
    for i, pos in enumerate(seq):
        row = col_row + 1 + i
        correct = get_correct(seq, i, n)
        n_ago  = seq[i - n] if i >= n else '-'
        answer = '-' if correct is None else ('はい' if correct else 'いいえ')
        bg = 'F2F6FF' if i % 2 == 0 else 'FFFFFF'

        values = [i + 1, pos, n_ago, answer, '', '']
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center()
            cell.border = thin_border()
            cell.font = Font(size=10)
            if col <= 4:
                cell.fill = PatternFill('solid', fgColor=bg)
            if col == 4 and answer == 'はい':
                cell.font = Font(size=10, bold=True, color='1565C0')
            if col == 4 and answer == '-':
                cell.font = Font(size=10, color='AAAAAA')
            if col == 3 and n_ago == '-':
                cell.font = Font(size=10, color='AAAAAA')
        ws.row_dimensions[row].height = 16

    # --- Accuracy footer ---
    footer_row = col_row + 1 + len(seq)
    ws.merge_cells(f'A{footer_row}:B{footer_row}')
    ws[f'A{footer_row}'].value = 'Accuracy:'
    ws[f'A{footer_row}'].font = Font(bold=True, size=10)
    ws[f'A{footer_row}'].alignment = center()
    for col in range(1, 7):
        ws.cell(row=footer_row, column=col).border = thick_bottom_border()
    ws[f'D{footer_row}'].value = f'/ {valid}'
    ws[f'D{footer_row}'].alignment = center()
    ws[f'F{footer_row}'].value = '%'
    ws[f'F{footer_row}'].alignment = center()
    ws.row_dimensions[footer_row].height = 16

    # Return next start row (one blank row gap)
    return footer_row + 2


# -----------------------------------------------
# Main
# -----------------------------------------------
def main():
    with open(SEQ_FILE, 'r', encoding='utf-8') as f:
        sequences = json.load(f)

    # Group keys by trial number
    trial_map = {}
    for key in sequences:
        m = re.search(r'trial(\d+)$', key)
        if m:
            t = int(m.group(1))
            trial_map.setdefault(t, []).append(key)

    # Sort conditions within each trial by CONDITION_ORDER
    for t in trial_map:
        trial_map[t].sort(key=lambda k: CONDITION_ORDER.index(
            re.sub(r'_trial\d+$', '', k)
        ) if re.sub(r'_trial\d+$', '', k) in CONDITION_ORDER else 99)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for trial_num in sorted(trial_map.keys()):
        ws = wb.create_sheet(title=f'Trial {trial_num}')

        # --- Sheet title ---
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'Trial {trial_num}  —  Answer Sheet'
        title_cell.font = Font(bold=True, size=13, color='FFFFFF')
        title_cell.fill = header_fill('2C3E50')
        title_cell.alignment = center()
        title_cell.border = thin_border()
        ws.row_dimensions[1].height = 22

        current_row = 3  # start below sheet title (one blank row gap)
        for key in trial_map[trial_num]:
            cond_key = re.sub(r'_trial\d+$', '', key)
            current_row = write_condition_block(ws, current_row, cond_key, sequences[key])

    wb.save(OUTPUT_FILE)
    print(f'Saved: {OUTPUT_FILE}  ({len(trial_map)} sheets)')

if __name__ == '__main__':
    main()